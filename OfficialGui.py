# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'test.ui'
#
# Created by: PyQt5 UI code generator 5.15.6
#
# Creator: Garrett Thompson
# Interface made with a combination of coding and pyqt Designer
#
# Imports for the code to work
import openpyxl
import datetime
from PyQt5 import QtCore, QtGui, QtWidgets
import API_Interaction
import ExcelBeautifier
import os


# Class of the main window and how I interact with it, started as a base window and fleshed it out to involve everything
# we needed.
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        # Controls window size
        MainWindow.resize(835, 461)
        # Controls the main window as a widget
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # Creates the run button visually
        self.RunButton = QtWidgets.QPushButton(self.centralwidget)
        self.RunButton.setGeometry(QtCore.QRect(650, 160, 111, 51))
        self.RunButton.setObjectName("RunButton")
        # Creates the drop down box visually
        self.DropDown = QtWidgets.QComboBox(self.centralwidget)
        self.DropDown.setGeometry(QtCore.QRect(30, 160, 581, 51))
        self.DropDown.setObjectName("DropDown")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        # Creates the title for the window
        self.TitleInfo = QtWidgets.QLabel(self.centralwidget)
        self.TitleInfo.setGeometry(QtCore.QRect(30, 40, 731, 111))
        self.TitleInfo.setTextFormat(QtCore.Qt.RichText)
        self.TitleInfo.setWordWrap(True)
        self.TitleInfo.setObjectName("TitleInfo")
        self.Mineslogo = QtWidgets.QLabel(self.centralwidget)
        self.Mineslogo.setGeometry(QtCore.QRect(600, 280, 161, 151))
        self.Mineslogo.setText("")
        # Creates the logo for the interface
        self.Mineslogo.setPixmap(QtGui.QPixmap(":/MinesLogo/Arthur Lakes Library Logo.jpg"))
        self.Mineslogo.setScaledContents(True)
        self.Mineslogo.setObjectName("Mineslogo")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        # How the run button interacts with actual code
        self.RunButton.clicked.connect(self.pressed)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        # Controls the UI translating up and down to avoid weird things happening with the dropdown box
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.RunButton.setText(_translate("MainWindow", "Run"))
        self.DropDown.setItemText(0, _translate("MainWindow", "1. ALL"))
        self.DropDown.setItemText(1, _translate("MainWindow", "2. Library.Mines.edu"))
        self.DropDown.setItemText(2, _translate("MainWindow", "4. LibGuides"))
        self.DropDown.setItemText(3, _translate("MainWindow", "5. Libcal"))
        self.DropDown.setItemText(4, _translate("MainWindow", "6. Libanswers"))
        self.DropDown.setItemText(5, _translate("MainWindow", "7. Libwaizard"))
        self.TitleInfo.setText(_translate("MainWindow",
                                          "<html><head/><body><p><span style=\" font-size:16pt;\">Welcome to the web stats automation software.  Select the site you would like to pull statistics for and press Run. Software created by: Garrett Thompson</span></p></body></html>"))

    def pressed(self):
        # Actual code needed for implementation of the drop-down box and Run button
        site_list = ['main-website_detail', 'libguides_detail', 'libcal_detail', 'libanswers_detail',
                     'libwizard_detail']
        choice_list = [1, 4, 5, 6, 7]
        choice = self.DropDown.currentText()
        choice = int(choice[:1])
        # Determins if we are running all at once or if we are doing specific sites based on user choice
        if choice != 1:
            if choice == 2:
                choice = 1
                site = site_list[0]
            elif choice == 4:
                site = site_list[1]
            elif choice == 5:
                site = site_list[2]
            elif choice == 6:
                site = site_list[3]
            elif choice == 7:
                site = site_list[4]
            # Controls the main path that the webstats are stored in
            default = ""
            output = default + site + '.csv'
            API_Interaction.page_titles(choice, output)
            # Calling outlinks function to get the data for outlinks(More decsription in API_interaction)
            API_Interaction.outlinks(choice, output)
            # Calling the refferring_urls function to get the data for referring urls(More decsription in API_interaction)
            API_Interaction.referring_urls(choice, output)
            # Calling the convert csv to excel file to ensure we are now making an excel file to edit later

            # Setting the date to last month to write to the excel file (dates are weird)
            today = datetime.date.today()
            first = today.replace(day=1)
            lastMonth = first - datetime.timedelta(days=1)
            d1 = lastMonth.strftime("%B %Y")
            d1 = d1[0:3] + " " + d1[d1.find(" ") + 1:]
            output_excel = output[:-4]
            output_excel += '.xlsx'
            # Creates a workbook by loading the previous webstats already stored (this was a pain)
            xfile = openpyxl.load_workbook(output_excel)
            xfile.create_sheet(d1)
            sheet = xfile[d1]
            # Spaces the columns and converts the csv to the excel sheet
            ExcelBeautifier.convert_csv_xlsx(output, sheet)
            ExcelBeautifier.space_columns(sheet)
            os.remove(output)
            xfile.save(output_excel)
        else:
            for i in range(0, len(site_list)):
                site = site_list[i]
                choice = choice_list[i]
                default = ""
                output = default + site + '.csv'
                API_Interaction.page_titles(choice, output)
                # Calling outlinks function to get the data for outlinks(More decsription in API_interaction)
                API_Interaction.outlinks(choice, output)
                # Calling the refferring_urls function to get the data for referring urls(More decsription in API_interaction)
                API_Interaction.referring_urls(choice, output)
                # Setting the date to last month to write to the excel file (dates are weird)
                today = datetime.date.today()
                first = today.replace(day=1)
                lastMonth = first - datetime.timedelta(days=1)
                d1 = lastMonth.strftime("%B %Y")
                d1 = d1[0:3] + " " + d1[d1.find(" ") + 1:]
                output_excel = output[:-4]
                output_excel += '.xlsx'
                # Creates a workbook by loading the previous webstats already stored (this was a pain)
                xfile = openpyxl.load_workbook(output_excel)
                xfile.create_sheet(d1)
                sheet = xfile[d1]
                # Spaces the columns and converts the csv to the excel sheet
                ExcelBeautifier.convert_csv_xlsx(output, sheet)
                ExcelBeautifier.space_columns(sheet)
                os.remove(output)
                xfile.save(output_excel)


import Designerpictures_rc

if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
