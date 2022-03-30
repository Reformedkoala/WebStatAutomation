# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'test.ui'
#
# Created by: PyQt5 UI code generator 5.15.6
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.
import openpyxl
from datetime import date
from PyQt5 import QtCore, QtGui, QtWidgets
import API_Interaction
import ExcelBeautifier
import os


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(835, 461)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.RunButton = QtWidgets.QPushButton(self.centralwidget)
        self.RunButton.setGeometry(QtCore.QRect(650, 160, 111, 51))
        self.RunButton.setObjectName("RunButton")
        self.DropDown = QtWidgets.QComboBox(self.centralwidget)
        self.DropDown.setGeometry(QtCore.QRect(30, 160, 581, 51))
        self.DropDown.setObjectName("DropDown")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.DropDown.addItem("")
        self.TitleInfo = QtWidgets.QLabel(self.centralwidget)
        self.TitleInfo.setGeometry(QtCore.QRect(30, 40, 731, 111))
        self.TitleInfo.setTextFormat(QtCore.Qt.RichText)
        self.TitleInfo.setWordWrap(True)
        self.TitleInfo.setObjectName("TitleInfo")
        self.Mineslogo = QtWidgets.QLabel(self.centralwidget)
        self.Mineslogo.setGeometry(QtCore.QRect(600, 280, 161, 151))
        self.Mineslogo.setText("")
        self.Mineslogo.setPixmap(QtGui.QPixmap(":/MinesLogo/Arthur Lakes Library Logo.jpg"))
        self.Mineslogo.setScaledContents(True)
        self.Mineslogo.setObjectName("Mineslogo")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.RunButton.clicked.connect(self.pressed)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.RunButton.setText(_translate("MainWindow", "Run"))
        self.DropDown.setItemText(0, _translate("MainWindow", "1. ALL"))
        self.DropDown.setItemText(1, _translate("MainWindow", "2. Library.Mines.edu"))
        self.DropDown.setItemText(2, _translate("MainWindow", "4. LibGuides"))
        self.DropDown.setItemText(3, _translate("MainWindow", "5. Libcal"))
        self.DropDown.setItemText(4, _translate("MainWindow", "6. Libanswers"))
        self.DropDown.setItemText(5, _translate("MainWindow", "7. Libwaizard"))
        self.TitleInfo.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Welcome to the web stats automation software.  Select the site you would like to pull statistics for and press Run. Software created by: Garrett Thompson</span></p></body></html>"))


    def pressed(self):
        site_list = ['main-website_detail', 'libguides_detail', 'libcal_detail', 'libanswers_detail', 'libwizard_detail']
        choice_list = [1, 4, 5, 6, 7]
        choice = self.DropDown.currentText()
        choice = int(choice[:1])
        if choice != 1:
            if choice == 2:
                choice = 1
                site = site_list[0]
            elif choice == 4:
                site = site_list[1]
            elif choice == 5:
                site = site_list[2]
            elif choice == 6:
                site = site_list[3]
            elif choice == 7:
                site = site_list[4]
            default = "Y:\LB\SharedSpace\Systems\WebStats\\"

            output = default + site +'.csv'
            API_Interaction.page_titles(choice, output)
            # Calling outlinks function to get the data for outlinks(More decsription in API_interaction)
            API_Interaction.outlinks(choice, output)
            # Calling the refferring_urls function to get the data for referring urls(More decsription in API_interaction)
            API_Interaction.referring_urls(choice, output)
            # Calling the convert csv to excel file to ensure we are now making an excel file to edit later

            today = date.today()
            d1 = today.strftime("%B %Y")
            d1 = d1[0:3] + " " + d1[d1.find(" ") + 1:]
            output_excel = output[:-4]
            output_excel += '.xlsx'
            xfile = openpyxl.load_workbook(output_excel)
            xfile.create_sheet(d1)
            sheet = xfile[d1]
            ExcelBeautifier.convert_csv_xlsx(output, sheet)
            ExcelBeautifier.space_columns(sheet)
            os.remove(output)
            xfile.save(output_excel)
        else:
            for i in range(0, len(site_list)):
                site = site_list[i]
                choice = choice_list[i]
                default = "Y:\LB\SharedSpace\Systems\WebStats\\"
                output = default + site + '.csv'
                API_Interaction.page_titles(choice, output)
                # Calling outlinks function to get the data for outlinks(More decsription in API_interaction)
                API_Interaction.outlinks(choice, output)
                # Calling the refferring_urls function to get the data for referring urls(More decsription in API_interaction)
                API_Interaction.referring_urls(choice, output)
                # Calling the convert csv to excel file to ensure we are now making an excel file to edit later
                today = date.today()
                d1 = today.strftime("%B %Y")
                d1 = d1[0:3] + " " + d1[d1.find(" ") + 1:]
                output_excel = output[:-4]
                output_excel += '.xlsx'
                xfile = openpyxl.load_workbook(output_excel)
                xfile.create_sheet(d1)
                sheet = xfile[d1]
                ExcelBeautifier.convert_csv_xlsx(output, sheet)
                ExcelBeautifier.space_columns(sheet)
                os.remove(output)
                xfile.save(output_excel)



import Designerpictures_rc
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
